import numpy as np
import pandas as pd
import openpyxl as op
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


def read_path_of_files():
    numberof_sheets= int(input("Enter Number of sheets To read And extract data: "))
    print("1. Enter the Paths of ", numberof_sheets , " Excel sheets")
    for i in range (1,numberof_sheets+1):
        print('Path of sheet', i, ':')
        filepath_ofSheet.append(input())



def read_files():
    
    for fpath  in (filepath_ofSheet):
        workbok1 = op.load_workbook(fpath)
        dff.append(pd.read_excel(fpath))
        namesheet = workbok1.sheetnames
        print(namesheet)
        allsheetnames.append(namesheet[0])
        wb_read.append(op.load_workbook(fpath))



def write_files_in_one():
    with pd.ExcelWriter('auto5sheets1.xlsx') as writer:
        for (df,a) in zip(dff, allsheetnames):
            df.to_excel(writer, sheet_name=a, index=False)



def header_in_new_sheets():
    MasterSheet= Workbook()
    wb_read=op.load_workbook("auto5sheets1.xlsx")
    wsheet= MasterSheet.active
    wsheet.title='output'
    sheets_read=wb_read.sheetnames
    for sheet in wb_read.sheetnames:
        rs=wb_read[sheet] 
        Maxrow=rs.max_row
        Maxcol=rs.max_column
        if sheet == allsheetnames[0]:
            for j in range(1,Maxcol+1):
                value=rs.cell(row=1, column=j).value
                wsheet.cell(row=1, column=j).value=value
        else:
            print("You are in now else block")
            maxColmaster = wsheet.max_column
            for j in range(7, Maxcol+1):
                value=rs.cell(row=1, column=j).value
                print(value)
                maxColmaster = maxColmaster+1
                wsheet.cell(row=1, column=maxColmaster).value=value
    MasterSheet.save('Openpyxl.xlsx')
# append_data();
# def append_data():
    # MasterSheet= Workbook()
    # wsheet= MasterSheet.active
    wb_read=op.load_workbook("auto5sheets1.xlsx")
    count=int(input("Enter how Many Data you want to read"))
    for i in range(1, count+1):
        print('Enter Details for the ', i, ' Data')
        FirstName=str(input('Enter First name '))
        Email=str(input('Enter email Id'))
        PS_No=int(input('Enter PS Number '))
        Datatoload=[]
        for sheet in wb_read.sheetnames:
            rs=wb_read[sheet]
            Maxrow=rs.max_row
            Maxcol=rs.max_column
            for i in range(2, Maxrow+1):
                if rs.cell(row=i, column=1).value == PS_No and rs.cell(row=i, column=2).value == FirstName and rs.cell(row=i, column=6).value == Email:
                    if sheet == allsheetnames[0]:
                        for j in range(1, Maxcol+1):
                            Datatoload.append(rs.cell(row=i, column=j).value)
                    else:
                        for j in range(7, Maxcol+1):
                            Datatoload.append(rs.cell(row=i, column=j).value)
                    df=pd.DataFrame(Datatoload)
                    df=df.T
        for r in dataframe_to_rows(df, index=False, header=False):
            wsheet.append(r)
        MasterSheet.save('Openpyxl.xlsx')     

# def append_data():
    
#     book = load_workbook('Openpyxl.xlsx')
#     writer = pd.ExcelWriter('Openpyxl.xlsx', engine='openpyxl')
#     writer.book = book
#     writer.sheets = {ws.title: ws for ws in book.worksheets}
#     wb_read=op.load_workbook("auto5sheets.xlsx")
#     count=int(input("Enter how Many Data you want to read"))
#     for i in range(1, count+1):
#         print('Enter Details for the ', i, ' Data')
#         FirstName=str(input('Enter First name '))
#         Email=str(input('Enter email Id'))
#         PS_No=int(input('Enter PS Number '))
#         Datatoload=[]
#         for sheet in wb_read.sheetnames:
#             rs=wb_read[sheet]
#             Maxrow=rs.max_row
#             Maxcol=rs.max_column
#             for i in range(2, Maxrow+1):
#                 if rs.cell(row=i, column=1).value == PS_No and rs.cell(row=i, column=2).value == FirstName and rs.cell(row=i, column=6).value == Email:
#                     if sheet == allsheetnames[0]:
#                         for j in range(1, Maxcol+1):
#                             Datatoload.append(rs.cell(row=i, column=j).value)
#                     else:
#                         for j in range(7, Maxcol+1):
#                             Datatoload.append(rs.cell(row=i, column=j).value)
#                     df=pd.DataFrame(Datatoload)
#                     df=df.T
#         # for r in dataframe_to_rows(df, index=False, header=False):
#         #     wsheet.append(r)
#         # MasterSheet.save('Openpyxl.xlsx')     
#     for sheetname in writer.sheets:
#         df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

#     writer.save()        
        
        
        
if __name__ == '__main__':
    
    filepath_ofSheet=[]
    allsheetnames = []
    dff=[]
    wb_read=[]
    read_path_of_files();
    read_files();
    write_files_in_one();
    
    # op=(input('you want to append or write in new file W/N'))
    #  op=='W' or op=='w':
    #     append_data();    
    # else:if
    header_in_new_sheets();
        
    
        
        
    
    
# D:\Python\python_project\a\quiz_1_grades.xlsx
# D:\Python\python_project\b\quiz_2_grades.xlsx
# D:\Python\python_project\c\quiz_3_grades.xlsx
# D:\Python\python_project\d\quiz_4_grades.xlsx
# D:\Python\python_project\e\quiz_5_grades.xlsx


# 99003700  Richard Bennett Male    22  richard.bennett@univ.edu
# 99003701  Timothy Parker  Female  25  timothy.parker@univ.edu
# 99003704  Michael Taylor  Male    20  michael.taylor@univ.edu
