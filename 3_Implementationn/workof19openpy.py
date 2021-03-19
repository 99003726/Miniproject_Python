    
import numpy as np
import pandas as pd
import openpyxl as op
from openpyxl import Workbook

filepath_ofSheet1=r"D:\Python\python_project\a\quiz_1_grades.xlsx"
filepath_ofSheet2=r"D:\Python\python_project\b\quiz_2_grades.xlsx"
filepath_ofSheet3=r"D:\Python\python_project\c\quiz_3_grades.xlsx"
filepath_ofSheet4=r"D:\Python\python_project\d\quiz_4_grades.xlsx"
filepath_ofSheet5=r"D:\Python\python_project\e\quiz_5_grades.xlsx"

dff1 = pd.read_excel(filepath_ofSheet1)
dff2 = pd.read_excel(filepath_ofSheet2)
dff3 = pd.read_excel(filepath_ofSheet3)
dff4 = pd.read_excel(filepath_ofSheet4)
dff5 = pd.read_excel(filepath_ofSheet5)


with pd.ExcelWriter('auto5sheets.xlsx') as writer:
    dff1.to_excel(writer, sheet_name='Sheet_Quiz_1', index=False)
    dff2.to_excel(writer, sheet_name='Sheet_Quiz_2', index=False)
    dff3.to_excel(writer, sheet_name='Sheet_Quiz_3', index=False)
    dff4.to_excel(writer, sheet_name='Sheet_Quiz_4', index=False)
    dff5.to_excel(writer, sheet_name='Sheet_Quiz_5', index=False)
    
# //////////////////=============================================Rough==============================================\\\\ 
# First_Name=str(input("Enter First Name"))
# Email=str(input("Enter Email Id"))
# PS_No =int(input("Enter your Ps. No."))

# wb_write=op.load_workbook("openpyxl.xlsx")
# wb_read=op.load_workbook("auto5sheets.xlsx")

# sheets_write=wb_write.sheetnames
# sheets_read=wb_read.sheetnames

# sh_read=wb_read.active
# sh_write=wb_write.active


# wb_write=op.load_workbook("openpyxl.xlsx")
# wb_read=op.load_workbook("auto5sheets.xlsx")

# sheets_write=wb_write.sheetnames
# sheets_read=wb_read.sheetnames
# for sheet in wb_read.sheetnames:
#     print(sheet)
#     ws=wb_read[sheet]
#     print(ws.cell(row=1, column=1).value)
    






# wb_write.save()

#/////////////// ================================================Roughhhhhhhhhh========================================================\\\\\\\\\
from openpyxl.utils.dataframe import dataframe_to_rows
MasterSheet= Workbook()
wb_read=op.load_workbook("auto5sheets.xlsx")
wsheet= MasterSheet.active
wsheet.title='output'


# a=int(input('how many data you want to extract'))
# count=0

# FirstName=str(input("Enter First Name"))
# Email=str(input("Enter Email Id"))
# PS_No =int(input("Enter your Ps. No."))

FirstName='Richard'
Email='richard.bennett@univ.edu'
PS_No=99003700


# sheets_write=wb_write.sheetnames



sheets_read=wb_read.sheetnames

for sheet in wb_read.sheetnames:
    rs=wb_read[sheet] 
    Maxrow=rs.max_row
    Maxcol=rs.max_column
    if rs == wb_read['Sheet_Quiz_1']:
        for j in range(1,Maxcol+1):
            value=rs.cell(row=1, column=j).value
            # print(value)
            wsheet.cell(row=1, column=j).value=value
    else:
        print("You are in now else block")
        maxColmaster = wsheet.max_column
        for j in range(7, Maxcol+1):
            value=rs.cell(row=1, column=j).value
            print(value)
            maxColmaster = maxColmaster+1
            wsheet.cell(row=1, column=maxColmaster).value=value
MasterSheet.save('Openpyxl.xlsx')
# for i in range(1, Maxrow+1):

Datatoload=[]
for sheet in wb_read.sheetnames:
    rs=wb_read[sheet] 
    Maxrow=rs.max_row
    Maxcol=rs.max_column
    for i in range(2, Maxrow+1):
        if rs.cell(row=i, column=1).value == PS_No and rs.cell(row=i, column=2).value == FirstName and rs.cell(row=i, column=6).value == Email:
            if rs == wb_read['Sheet_Quiz_1']:
                #   print(rs.cell(row=i, column=1).value)
                #   print(rs.cell(row=i, column=2).value)
                #   print(rs.cell(row=i, column=6).value)
                for j in range(1, Maxcol+1):
                    Datatoload.append(rs.cell(row=i, column=j).value)
                    
                    #   print(value)
                    #   Wsheet.cell(row=i, column=j).value=value

                            

            else:
                # print(rs.cell(row=i, column=2).value)
                # print(rs.cell(row=i, column=6).value)
                for j in range(7, Maxcol+1):
                    Datatoload.append(rs.cell(row=i, column=j).value)
                    #   print(value)
                    #   Wsheet.cell(row=i, column=j).value=value
df=pd.DataFrame(Datatoload)  
df=df.T             
for r in dataframe_to_rows(df, index=False, header=False):
    wsheet.append(r)
            
        

MasterSheet.save('Openpyxl.xlsx')       
